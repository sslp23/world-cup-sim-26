"""
Simulates the 2026 FIFA World Cup knockout stages and adds bracket tabs to
simulation/output/wc_26_ordered_logit.xlsx.

Run from the project root:
    python -m simulation.ordered_logit.playoff
"""

import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.ordered_logit.model import OrderedLogitPredictor
from simulation.dataset import build as build_dataset

XLSX_PATH  = 'simulation/output/wc_26_ordered_logit.xlsx'
WC26_START = pd.Timestamp('2026-06-11')

GROUPS = {
    'A': ['Mexico', 'South Africa', 'South Korea', 'Czech Republic'],
    'B': ['Canada', 'Bosnia and Herzegovina', 'Qatar', 'Switzerland'],
    'C': ['Brazil', 'Morocco', 'Haiti', 'Scotland'],
    'D': ['United States', 'Paraguay', 'Australia', 'Turkey'],
    'E': ['Germany', 'Curaçao', 'Ivory Coast', 'Ecuador'],
    'F': ['Netherlands', 'Japan', 'Sweden', 'Tunisia'],
    'G': ['Belgium', 'Egypt', 'Iran', 'New Zealand'],
    'H': ['Spain', 'Cape Verde', 'Saudi Arabia', 'Uruguay'],
    'I': ['France', 'Senegal', 'Iraq', 'Norway'],
    'J': ['Argentina', 'Algeria', 'Austria', 'Jordan'],
    'K': ['Portugal', 'DR Congo', 'Uzbekistan', 'Colombia'],
    'L': ['England', 'Croatia', 'Ghana', 'Panama'],
}

R32 = [
    (73, '2A',    '2B'),
    (74, '1E',    '3ABCDF'),
    (75, '1F',    '2C'),
    (76, '1C',    '2F'),
    (77, '1I',    '3CDFGH'),
    (78, '2E',    '2I'),
    (79, '1A',    '3CEFHI'),
    (80, '1L',    '3EHIJK'),
    (81, '1D',    '3BEFIJ'),
    (82, '1G',    '3AEHIJ'),
    (83, '2K',    '2L'),
    (84, '1H',    '2J'),
    (85, '1B',    '3EFGIJ'),
    (86, '1J',    '2H'),
    (87, '1K',    '3DEIJL'),
    (88, '2D',    '2G'),
]

R16 = [(89,74,77),(90,73,75),(91,76,78),(92,79,80),
       (93,83,84),(94,81,82),(95,86,88),(96,85,87)]
QF  = [(97,89,90),(98,93,94),(99,91,92),(100,95,96)]
SF  = [(101,97,98),(102,99,100)]

DARK_BLUE = 'FF1F4E79'
MID_BLUE  = 'FF2E75B6'
GREEN     = 'FFE2EFDA'
GOLD      = 'FFFFF2CC'
SILVER    = 'FFF2F2F2'
WHITE     = 'FFFFFFFF'

def _border():
    s = Side(style='thin', color='FFBFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


# ── Group standings ────────────────────────────────────────────────────────────

def _win_prob_sum(team, pred_df):
    total = 0.0
    for _, row in pred_df.iterrows():
        if row['Home'] == team:   total += row['P(HW)']
        elif row['Away'] == team: total += row['P(AW)']
    return total


def compute_standings(pred_df):
    team_pts  = {t: 0 for g in GROUPS.values() for t in g}
    team_wins = {t: 0 for g in GROUPS.values() for t in g}
    for _, row in pred_df.iterrows():
        h, a, pred = row['Home'], row['Away'], row['_pred_raw']
        if pred == 'home_win':
            team_pts[h] += 3; team_wins[h] += 1
        elif pred == 'draw':
            team_pts[h] += 1; team_pts[a] += 1
        else:
            team_pts[a] += 3; team_wins[a] += 1
    standings = {}
    for grp, teams in GROUPS.items():
        ranked = sorted(teams, key=lambda t: (-team_pts.get(t,0), -team_wins.get(t,0)))
        standings[grp] = [(t, team_pts.get(t,0), team_wins.get(t,0),
                           _win_prob_sum(t, pred_df)) for t in ranked]
    return standings


def pick_best_thirds(standings):
    thirds = []
    for grp, rows in standings.items():
        t, pts, wins, wps = rows[2]
        thirds.append((pts, wins, wps, t, grp))
    thirds.sort(key=lambda x: (-x[0], -x[1], -x[2]))
    return thirds[:8]


def allocate_thirds(best_thirds):
    COL_TO_MATCH = {
        '1A vs': 79, '1B vs': 85, '1D vs': 81, '1E vs': 74,
        '1G vs': 82, '1I vs': 77, '1K vs': 87, '1L vs': 80,
    }
    combo_df = pd.read_csv('simulation/third_place_combinations.csv')
    grp_cols = [c for c in combo_df.columns if 'groupsvte' in c]
    team_by_grp = {grp: team for _, _, _, team, grp in best_thirds}
    qual_groups = frozenset(team_by_grp.keys())

    match_row = None
    for _, row in combo_df.iterrows():
        row_groups = frozenset(str(v) for v in row[grp_cols] if pd.notna(v))
        if row_groups == qual_groups:
            match_row = row; break
    if match_row is None:
        raise ValueError(f'No combination found for groups: {sorted(qual_groups)}')

    assignment = {}
    for col, match_id in COL_TO_MATCH.items():
        val = str(match_row[col]).strip()
        grp = val.lstrip('3')
        if grp in team_by_grp:
            assignment[match_id] = team_by_grp[grp]
    return assignment


# ── Team profiles ──────────────────────────────────────────────────────────────

def build_profiles():
    """Extract per-team rating features from wc_26_data.csv."""
    wc_df = pd.read_csv('data/wc_26_data.csv')
    profiles = {}
    for _, row in wc_df.iterrows():
        for side in ('home', 'away'):
            team = row[f'{side}_team']
            if team not in profiles:
                profiles[team] = {
                    'elo':         row.get(f'{side}_elo', 1500),
                    'points':      row.get(f'points_{side}', 1400),
                    'mv_top20_sum': row.get(f'{side}_mv_top20_sum', np.nan),
                }
    return profiles


# ── Knockout prediction ────────────────────────────────────────────────────────

def predict_ko(model, team_a, team_b, profiles):
    """
    Predict a knockout match. Ordered Logit only needs rating features.
    Draws split proportionally by relative win probability.
    """
    pa = profiles.get(team_a, {})
    pb = profiles.get(team_b, {})

    row = pd.Series({
        'elo_diff':           pa.get('elo', 1500)    - pb.get('elo', 1500),
        'points_dif':         pa.get('points', 1400) - pb.get('points', 1400),
        'home_mv_top20_sum':  pa.get('mv_top20_sum', np.nan),
        'away_mv_top20_sum':  pb.get('mv_top20_sum', np.nan),
    })

    probs = model.predict_proba_row(row)
    p_hw, p_d, p_aw = probs['home_win'], probs['draw'], probs['away_win']

    p_a_win = p_hw + p_d * 0.5
    p_b_win = p_aw + p_d * 0.5

    winner = team_a if p_a_win >= p_b_win else team_b
    return winner, round(p_a_win, 3), round(p_b_win, 3)


# ── Bracket simulation ─────────────────────────────────────────────────────────

def simulate_bracket(standings, thirds_assign, model, profiles):
    slot_to_team = {}
    for grp, rows in standings.items():
        slot_to_team[f'1{grp}'] = rows[0][0]
        slot_to_team[f'2{grp}'] = rows[1][0]
    for mid, team in thirds_assign.items():
        for m_id, slotA, slotB in R32:
            if m_id == mid:
                key = f'r32_{mid}_B' if slotB.startswith('3') else f'r32_{mid}_A'
                slot_to_team[key] = team

    results = {}

    def resolve(slot, mid=None):
        if slot.startswith(('1', '2')):
            return slot_to_team[slot]
        if slot.startswith('3'):
            return slot_to_team.get(f'r32_{mid}_B', slot_to_team.get(f'r32_{mid}_A', '?'))
        return results[int(slot)]['winner']

    for mid, sA, sB in R32:
        tA, tB = resolve(sA, mid), resolve(sB, mid)
        w, pA, pB = predict_ko(model, tA, tB, profiles)
        results[mid] = {'home': tA, 'away': tB, 'winner': w,
                        'p_home': pA, 'p_away': pB, 'round': 'R32'}

    for mid, sA, sB in R16:
        tA, tB = results[sA]['winner'], results[sB]['winner']
        w, pA, pB = predict_ko(model, tA, tB, profiles)
        results[mid] = {'home': tA, 'away': tB, 'winner': w,
                        'p_home': pA, 'p_away': pB, 'round': 'R16'}

    for mid, sA, sB in QF:
        tA, tB = results[sA]['winner'], results[sB]['winner']
        w, pA, pB = predict_ko(model, tA, tB, profiles)
        results[mid] = {'home': tA, 'away': tB, 'winner': w,
                        'p_home': pA, 'p_away': pB, 'round': 'QF'}

    sf_losers = {}
    for mid, sA, sB in SF:
        tA, tB = results[sA]['winner'], results[sB]['winner']
        w, pA, pB = predict_ko(model, tA, tB, profiles)
        sf_losers[mid] = tB if w == tA else tA
        results[mid] = {'home': tA, 'away': tB, 'winner': w,
                        'p_home': pA, 'p_away': pB, 'round': 'SF'}

    t3a, t3b = sf_losers[SF[0][0]], sf_losers[SF[1][0]]
    w, pA, pB = predict_ko(model, t3a, t3b, profiles)
    results[103] = {'home': t3a, 'away': t3b, 'winner': w,
                    'p_home': pA, 'p_away': pB, 'round': '3rd'}

    fA, fB = results[SF[0][0]]['winner'], results[SF[1][0]]['winner']
    w, pA, pB = predict_ko(model, fA, fB, profiles)
    results[104] = {'home': fA, 'away': fB, 'winner': w,
                    'p_home': pA, 'p_away': pB, 'round': 'Final'}

    return results


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, merge_to=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _border()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return cell


def _data(ws, row, col, val, bold=False, bg=WHITE, center=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
    cell.border    = _border()
    return cell


def _write_round(ws, label, matches_info, results):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24
    _hdr(ws, 1, 1, label, size=13, merge_to=7)
    for ci, (col, w) in enumerate(zip(
        ['Match','Date','Team A','P(A)','P(B)','Team B','Winner'],
        [8, 10, 24, 8, 8, 24, 24]
    ), 1):
        _hdr(ws, 3, ci, col, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, (mid, date_lbl) in enumerate(matches_info, start=4):
        r  = results[mid]
        wA = r['winner'] == r['home']
        bg = SILVER if ri % 2 == 0 else WHITE
        _data(ws, ri, 1, f'M{mid}', bg=bg)
        _data(ws, ri, 2, date_lbl, bg=bg)
        _data(ws, ri, 3, r['home'], bold=wA,  bg=GREEN if wA  else bg, center=False)
        _data(ws, ri, 4, f"{r['p_home']:.1%}", bg=bg)
        _data(ws, ri, 5, f"{r['p_away']:.1%}", bg=bg)
        _data(ws, ri, 6, r['away'], bold=not wA, bg=GREEN if not wA else bg, center=False)
        _data(ws, ri, 7, r['winner'], bold=True, bg=GOLD)
    ws.freeze_panes = 'A4'


def _write_final(ws, results):
    ws.sheet_view.showGridLines = False
    _hdr(ws, 1, 1, 'FINAL & 3RD PLACE PLAYOFF', size=13, merge_to=7)
    for ci, (col, w) in enumerate(zip(
        ['Match','Date','Team A','P(A)','P(B)','Team B','Winner'],
        [8, 10, 24, 8, 8, 24, 24]
    ), 1):
        _hdr(ws, 3, ci, col, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, (mid, date_lbl) in enumerate([(103,'Jul 18'),(104,'Jul 19')], start=4):
        r  = results[mid]
        wA = r['winner'] == r['home']
        bg = SILVER if ri % 2 == 0 else WHITE
        _data(ws, ri, 1, f'M{mid}', bg=bg)
        _data(ws, ri, 2, date_lbl, bg=bg)
        _data(ws, ri, 3, r['home'], bold=wA,  bg=GREEN if wA  else bg, center=False)
        _data(ws, ri, 4, f"{r['p_home']:.1%}", bg=bg)
        _data(ws, ri, 5, f"{r['p_away']:.1%}", bg=bg)
        _data(ws, ri, 6, r['away'], bold=not wA, bg=GREEN if not wA else bg, center=False)
        _data(ws, ri, 7, r['winner'], bold=True, bg=GOLD)
    champion = results[104]['winner']
    ws.row_dimensions[7].height = 28
    c = ws.cell(row=7, column=1, value=f'PREDICTED CHAMPION:  {champion}')
    c.font      = Font(bold=True, size=14, color=DARK_BLUE)
    c.fill      = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = _border()
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    print('Loading group stage predictions...')
    xl  = pd.ExcelFile(XLSX_PATH)
    df  = xl.parse(xl.sheet_names[0])
    df  = df[df['Home'].notna() & df['Away'].notna()].copy()
    df['_pred_raw'] = df.apply(
        lambda r: 'home_win' if r['Predicted'] == r['Home']
                  else ('away_win' if r['Predicted'] == r['Away'] else 'draw'), axis=1)

    print('Computing group standings...')
    standings = compute_standings(df)
    for grp, rows in standings.items():
        print(f"  Group {grp}: " + ", ".join(f"{t}({p}pts)" for t,p,w,_ in rows))

    best_thirds = pick_best_thirds(standings)
    print(f'\n8 best 3rd-place teams:')
    for pts, wins, wps, team, grp in best_thirds:
        print(f'  {team} (Group {grp}) — {pts}pts, {wins}W')

    thirds_assign = allocate_thirds(best_thirds)
    print(f'\n3rd-place slot assignments:')
    for mid, team in sorted(thirds_assign.items()):
        print(f'  Match {mid}: {team}')

    print('\nTraining Ordered Logit...')
    full_df  = build_dataset()
    train_df = full_df[full_df['date'] < WC26_START].reset_index(drop=True)
    model    = OrderedLogitPredictor()
    model.fit(train_df)

    print('Building team profiles...')
    profiles = build_profiles()

    print('Simulating bracket...')
    results = simulate_bracket(standings, thirds_assign, model, profiles)

    print('\nWriting Excel tabs...')
    wb = load_workbook(XLSX_PATH)
    for name in ['Round of 32','Round of 16','Quarter-finals','Semi-finals','Final']:
        if name in wb.sheetnames:
            del wb[name]

    r32_dates = ['Jun 28','Jun 29','Jun 29','Jun 29','Jun 30','Jun 30',
                 'Jun 30','Jul 1', 'Jul 1', 'Jul 1', 'Jul 2', 'Jul 2',
                 'Jul 2', 'Jul 3', 'Jul 3', 'Jul 3']
    r16_dates = ['Jul 4','Jul 4','Jul 5','Jul 5','Jul 6','Jul 6','Jul 7','Jul 7']
    qf_dates  = ['Jul 9','Jul 10','Jul 11','Jul 11']
    sf_dates  = ['Jul 14','Jul 15']

    _write_round(wb.create_sheet('Round of 32'), 'ROUND OF 32',
                 [(mid, d) for (mid,_,__), d in zip(R32, r32_dates)], results)
    _write_round(wb.create_sheet('Round of 16'), 'ROUND OF 16',
                 [(mid, d) for (mid,_,__), d in zip(R16, r16_dates)], results)
    _write_round(wb.create_sheet('Quarter-finals'), 'QUARTER-FINALS',
                 [(mid, d) for (mid,_,__), d in zip(QF, qf_dates)], results)
    _write_round(wb.create_sheet('Semi-finals'), 'SEMI-FINALS',
                 [(mid, d) for (mid,_,__), d in zip(SF, sf_dates)], results)
    _write_final(wb.create_sheet('Final'), results)

    wb.save(XLSX_PATH)
    print(f'\nSaved: {XLSX_PATH}')
    print(f'PREDICTED CHAMPION: {results[104]["winner"]}')
    print(f'3rd place:          {results[103]["winner"]}')


if __name__ == '__main__':
    run()
