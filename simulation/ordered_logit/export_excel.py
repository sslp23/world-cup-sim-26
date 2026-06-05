"""
Exports simulation predictions to an Excel file.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = 'FF1F4E79'
HW_COLOR  = 'FFD9E1F2'
DR_COLOR  = 'FFFFF2CC'
AW_COLOR  = 'FFFCE4D6'
GRAY      = 'FFF2F2F2'


def _border():
    s = Side(style='thin', color='FFBFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _pred_fill(pred_raw):
    if pred_raw == 'home_win': return PatternFill('solid', fgColor=HW_COLOR)
    if pred_raw == 'draw':     return PatternFill('solid', fgColor=DR_COLOR)
    return PatternFill('solid', fgColor=AW_COLOR)


def save(predictions: list[dict], out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'WC 2026 Group Stage'

    cols   = ['Date', 'Home', 'Away', 'P(HW)', 'P(D)', 'P(AW)', 'Predicted', 'Confidence']
    widths = [12, 26, 26, 8, 8, 8, 26, 13]
    keys   = ['date', 'home', 'away', 'p_home_win', 'p_draw', 'p_away_win', 'predicted', 'confidence']

    hfont = Font(bold=True, color='FFFFFFFF', size=11)
    hfill = PatternFill('solid', fgColor=DARK_BLUE)
    halign = Alignment(horizontal='center')

    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, pred in enumerate(predictions, 2):
        pfill = _pred_fill(pred['_pred_raw'])
        for ci, key in enumerate(keys, 1):
            c = ws.cell(row=ri, column=ci, value=pred[key])
            c.border    = _border()
            c.alignment = Alignment(horizontal='left' if ci <= 3 else 'center')
            if key in ('predicted', 'confidence'):
                c.fill = pfill

    # Summary block
    sep  = len(predictions) + 3
    hw   = sum(1 for p in predictions if p['_pred_raw'] == 'home_win')
    dr   = sum(1 for p in predictions if p['_pred_raw'] == 'draw')
    aw   = sum(1 for p in predictions if p['_pred_raw'] == 'away_win')
    bold = Font(bold=True)
    gfill = PatternFill('solid', fgColor=GRAY)

    ws.cell(row=sep - 1, column=1, value='SUMMARY').font = Font(bold=True, color=DARK_BLUE, size=11)
    for r, label, val in [
        (sep,     'Total matches',       len(predictions)),
        (sep + 1, 'Predicted home wins', hw),
        (sep + 2, 'Predicted draws',     dr),
        (sep + 3, 'Predicted away wins', aw),
    ]:
        c = ws.cell(row=r, column=1, value=label)
        c.font, c.fill = bold, gfill
        ws.cell(row=r, column=2, value=val)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'Saved: {out_path}')
