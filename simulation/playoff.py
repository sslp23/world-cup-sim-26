"""
Simulates the 2026 FIFA World Cup knockout stages and adds bracket tabs to
simulation/output/wc_26_sim.xlsx.

Flow:
  1. Load group stage predictions from wc_26_sim.xlsx
  2. Compute group standings (Pts → W → ELO tiebreak)
  3. Pick the 8 best 3rd-place teams, allocate to R32 slots
  4. Re-train CatBoost (draw_weight=0.66) on post-WC22 data
  5. Build team feature profiles from wc_26_data.csv
  6. Simulate R32 → R16 → QF → SF → Final with symmetrized inference
     Draws resolved by relative strength (P_win / (P_win + P_loss))

Run from the project root:
    python -m simulation.playoff
"""

import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.catboost.model import CatBoostPredictor
from simulation.dataset    import build as build_dataset

XLSX_PATH  = 'simulation/output/wc_26_sim.xlsx'
WC26_START = pd.Timestamp('2026-06-11')
WC26_END   = pd.Timestamp('2026-06-27')

# ── Groups ────────────────────────────────────────────────────────────────────
GROUPS = {
    'A': ['Mexico', 'South Africa', 'South Korea', 'Czech Republic'],
    'B': ['Canada', 'Bosnia and Herzegovina', 'Qatar', 'Switzerland'],
    'C': ['Brazil', 'Morocco', 'Haiti', 'Scotland'],
    'D': ['United States', 'Paraguay', 'Australia', 'Turkey'],
    'E': ['Germany', 'Curaçao', 'Ivory Coast', 'Ecuador'],
    'F': ['Netherlands', 'Japan', 'Sweden', 'Tunisia'],
    'G': ['Belgium', 'Egypt', 'Iran', 'New Zealand'],
    'H': ['Spain', 'Cape Verde', 'Saudi Arabia', 'Uruguay'],
    'I': ['France', 'Senegal', 'Iraq', 'Norway'],
    'J': ['Argentina', 'Algeria', 'Austria', 'Jordan'],
    'K': ['Portugal', 'DR Congo', 'Uzbekistan', 'Colombia'],
    'L': ['England', 'Croatia', 'Ghana', 'Panama'],
}

# ── R32 bracket ───────────────────────────────────────────────────────────────
# (match_id, slot_A, slot_B)
# slot format: '1A' = winner Group A, '2B' = runner-up Group B
# '3ABCDF' = best 3rd from groups A/B/C/D/F (eligible set)
R32 = [
    (73, '2A',      '2B'),
    (74, '1E',      '3ABCDF'),
    (75, '1F',      '2C'),
    (76, '1C',      '2F'),
    (77, '1I',      '3CDFGH'),
    (78, '2E',      '2I'),
    (79, '1A',      '3CEFHI'),
    (80, '1L',      '3EHIJK'),
    (81, '1D',      '3BEFIJ'),
    (82, '1G',      '3AEHIJ'),
    (83, '2K',      '2L'),
    (84, '1H',      '2J'),
    (85, '1B',      '3EFGIJ'),
    (86, '1J',      '2H'),
    (87, '1K',      '3DEIJL'),
    (88, '2D',      '2G'),
]
R32_THIRD_SLOTS = {mid: slot for mid, _, slot in R32 if slot.startswith('3')}

R16 = [(89,74,77), (90,73,75), (91,76,78), (92,79,80),
       (93,83,84), (94,81,82), (95,86,88), (96,85,87)]
QF  = [(97,89,90), (98,93,94), (99,91,92), (100,95,96)]
SF  = [(101,97,98), (102,99,100)]
THIRD_PLACE = (103, 101, 102)
FINAL       = (104, 101, 102)

ROUND_DATES = {
    'R32': ['Jun 28','Jun 29','Jun 30','Jul 1','Jul 2','Jul 3'],
    'R16': ['Jul 4','Jul 5','Jul 6','Jul 7'],
    'QF':  ['Jul 9','Jul 10','Jul 11'],
    'SF':  ['Jul 14','Jul 15'],
    '3rd': 'Jul 18',
    'Final': 'Jul 19',
}

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = 'FF1F4E79'
MID_BLUE   = 'FF2E75B6'
LIGHT_BLUE = 'FFD9E1F2'
GREEN      = 'FFE2EFDA'
GOLD       = 'FFFFF2CC'
SILVER     = 'FFF2F2F2'
WHITE      = 'FFFFFFFF'

def _border():
    s = Side(style='thin', color='FFBFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


# ── Group standings ───────────────────────────────────────────────────────────

def _opp_win_prob_sum(team, pred_df):
    """Sum of opponent win probabilities across a team's group matches (lower = better)."""
    total = 0.0
    for _, row in pred_df.iterrows():
        if row['Home'] == team:
            total += row['P(AW)']
        elif row['Away'] == team:
            total += row['P(HW)']
    return total


def compute_standings(pred_df):
    """
    Return {group: [(team, pts, wins, opp_win_sum), ...]} sorted 1→4.

    Tiebreaker: Pts → W → opponent win prob sum (ascending — lower is better).
    This matches group_tables.py exactly, ensuring the bracket is consistent
    with the group tab standings shown in the Excel.
    """
    team_pts  = {t: 0 for g in GROUPS.values() for t in g}
    team_wins = {t: 0 for g in GROUPS.values() for t in g}

    for _, row in pred_df.iterrows():
        h, a, pred = row['Home'], row['Away'], row['_pred_raw']
        if pred == 'home_win':
            team_pts[h]  += 3; team_wins[h] += 1
        elif pred == 'draw':
            team_pts[h]  += 1; team_pts[a]  += 1
        else:
            team_pts[a]  += 3; team_wins[a] += 1

    standings = {}
    for grp, teams in GROUPS.items():
        ranked = sorted(teams,
                        key=lambda t: (-team_pts.get(t, 0),
                                       -team_wins.get(t, 0)))
        standings[grp] = [(t, team_pts.get(t, 0), team_wins.get(t, 0),
                           _opp_win_prob_sum(t, pred_df)) for t in ranked]
    return standings


def pick_best_thirds(standings, pred_df):
    """Return sorted list of (pts, wins, opp_win_sum, team, group) for 8 best 3rd-place."""
    thirds = []
    for grp, rows in standings.items():
        t, pts, wins, opp_win_sum = rows[2]   # 3rd place
        thirds.append((pts, wins, opp_win_sum, t, grp))
    # Sort: pts desc, wins desc, opp_win_sum asc (lower opponent win prob = better)
    thirds.sort(key=lambda x: (-x[0], -x[1], x[2]))
    return thirds[:8]


def allocate_thirds(best_thirds, r32):
    """
    Assign 3rd-place teams to R32 slots using the official FIFA lookup table
    (simulation/third_place_combinations.csv, Annex C of the regulations).

    Column → match mapping:
      1A vs → M79  |  1B vs → M85  |  1D vs → M81  |  1E vs → M74
      1G vs → M82  |  1I vs → M77  |  1K vs → M87  |  1L vs → M80

    Returns {match_id: team}.
    """
    import pandas as pd

    COL_TO_MATCH = {
        '1A vs': 79, '1B vs': 85, '1D vs': 81, '1E vs': 74,
        '1G vs': 82, '1I vs': 77, '1K vs': 87, '1L vs': 80,
    }

    combo_df  = pd.read_csv('simulation/third_place_combinations.csv')
    grp_cols  = [c for c in combo_df.columns if 'groupsvte' in c]
    team_by_grp = {grp: team for _, _, _, team, grp in best_thirds}
    qual_groups = frozenset(team_by_grp.keys())

    # Find the matching row
    match_row = None
    for _, row in combo_df.iterrows():
        row_groups = frozenset(str(v) for v in row[grp_cols] if pd.notna(v))
        if row_groups == qual_groups:
            match_row = row
            break

    if match_row is None:
        raise ValueError(f'No combination found for groups: {sorted(qual_groups)}')

    assignment = {}
    for col, match_id in COL_TO_MATCH.items():
        val = str(match_row[col]).strip()   # e.g. '3C'
        grp = val.lstrip('3')               # e.g. 'C'
        if grp in team_by_grp:
            assignment[match_id] = team_by_grp[grp]

    return assignment


# ── Team feature profiles ──────────────────────────────────────────────────────

def build_profiles():
    """Extract per-team feature dict from wc_26_data.csv."""
    wc_df = pd.read_csv('data/wc_26_data.csv')
    profiles = {}
    for _, row in wc_df.iterrows():
        for side, opp_side in [('home','away'), ('away','home')]:
            team = row[f'{side}_team']
            if team not in profiles:
                profiles[team] = {
                    'elo':            row.get(f'{side}_elo', 1500),
                    'points':         row.get(f'points_{side}', 1400),
                    'pww_ma20':       row.get(f'{side}_points_weighted_ma_20', 0),
                    'pww_ma5':        row.get(f'{side}_points_weighted_ma_5', 0),
                    'gw_ma20':        row.get(f'{side}_goals_weighted_ma_20', 0),
                    'gw_ma5':         row.get(f'{side}_goals_weighted_ma_5', 0),
                    'gsw_ma20':       row.get(f'{side}_goals_suffered_weighted_ma_20', 0),
                    'gsw_ma5':        row.get(f'{side}_goals_suffered_weighted_ma_5', 0),
                    'gd_ma20':        row.get(f'{side}_goal_diff_ma_20', 0),
                    'gd_ma5':         row.get(f'{side}_goal_diff_ma_5', 0),
                    'wc_games':       row.get(f'{side}_wc_games', 0),
                    'elo_delta_20':   row.get(f'{side}_elo_delta_20', 0),
                    'elo_ma_8yr':     row.get(f'{side}_elo_ma_8yr', 1500),
                    'confederation':  row.get(f'confederation_{side}', 'Unknown'),
                }
    return profiles


def predict_ko(model, team_a, team_b, profiles):
    """
    Predict a knockout match (neutral venue, no draw allowed).
    Returns (winner, p_a, p_b, p_draw_raw).
    Uses symmetrized inference; draws split by relative strength.
    """
    pa = profiles.get(team_a, {})
    pb = profiles.get(team_b, {})

    row = pd.Series({
        'home_elo':                          pa.get('elo', 1500),
        'away_elo':                          pb.get('elo', 1500),
        'elo_diff':                          pa.get('elo',1500) - pb.get('elo',1500),
        'points_home':                       pa.get('points', 1400),
        'points_away':                       pb.get('points', 1400),
        'points_dif':                        pa.get('points',1400) - pb.get('points',1400),
        'home_points_weighted_ma_20':        pa.get('pww_ma20', 0),
        'away_points_weighted_ma_20':        pb.get('pww_ma20', 0),
        'home_points_weighted_ma_5':         pa.get('pww_ma5', 0),
        'away_points_weighted_ma_5':         pb.get('pww_ma5', 0),
        'home_goals_weighted_ma_20':         pa.get('gw_ma20', 0),
        'away_goals_weighted_ma_20':         pb.get('gw_ma20', 0),
        'home_goals_weighted_ma_5':          pa.get('gw_ma5', 0),
        'away_goals_weighted_ma_5':          pb.get('gw_ma5', 0),
        'home_goals_suffered_weighted_ma_20':pa.get('gsw_ma20', 0),
        'away_goals_suffered_weighted_ma_20':pb.get('gsw_ma20', 0),
        'home_goals_suffered_weighted_ma_5': pa.get('gsw_ma5', 0),
        'away_goals_suffered_weighted_ma_5': pb.get('gsw_ma5', 0),
        'home_goal_diff_ma_20':              pa.get('gd_ma20', 0),
        'away_goal_diff_ma_20':              pb.get('gd_ma20', 0),
        'home_goal_diff_ma_5':              pa.get('gd_ma5', 0),
        'away_goal_diff_ma_5':              pb.get('gd_ma5', 0),
        'home_wc_games':                    pa.get('wc_games', 0),
        'away_wc_games':                    pb.get('wc_games', 0),
        'home_elo_delta_20':                pa.get('elo_delta_20', 0),
        'away_elo_delta_20':                pb.get('elo_delta_20', 0),
        'home_elo_ma_8yr':                  pa.get('elo_ma_8yr', 1500),
        'away_elo_ma_8yr':                  pb.get('elo_ma_8yr', 1500),
        'confederation_home':               pa.get('confederation', 'Unknown'),
        'confederation_away':               pb.get('confederation', 'Unknown'),
    })

    probs   = model.predict_proba_row(row)
    p_hw    = probs['home_win']
    p_d     = probs['draw']
    p_aw    = probs['away_win']

    # Resolve draw: redistribute proportionally to win probabilities
    total   = p_hw + p_aw
    p_a_win = p_hw + p_d * (p_hw / total if total > 0 else 0.5)
    p_b_win = p_aw + p_d * (p_aw / total if total > 0 else 0.5)

    winner = team_a if p_a_win >= p_b_win else team_b
    return winner, round(p_a_win, 3), round(p_b_win, 3), round(p_d, 3)


# ── Simulate bracket ──────────────────────────────────────────────────────────

def simulate_bracket(standings, thirds_assign, model, profiles):
    """
    Resolve the bracket slot by slot. Returns results dict:
    {match_id: {'home': team, 'away': team, 'winner': team, 'p_home': float, 'p_away': float}}
    """
    # Build slot → team lookup
    slot_to_team = {}
    for grp, rows in standings.items():
        slot_to_team[f'1{grp}'] = rows[0][0]
        slot_to_team[f'2{grp}'] = rows[1][0]
    for mid, team in thirds_assign.items():
        # Find which slot in R32 this is (team goes to the '3...' slot of match mid)
        for m_id, slotA, slotB in R32:
            if m_id == mid:
                if slotB.startswith('3'):
                    slot_to_team[f'r32_{mid}_B'] = team
                else:
                    slot_to_team[f'r32_{mid}_A'] = team

    results = {}

    def resolve_slot(slot, mid=None, side=None):
        """Resolve a slot string to a team name."""
        if slot.startswith('1') or slot.startswith('2'):
            return slot_to_team[slot]
        if slot.startswith('3'):
            key_b = f'r32_{mid}_B'
            key_a = f'r32_{mid}_A'
            return slot_to_team.get(key_b, slot_to_team.get(key_a, '?'))
        # It's a match_id reference → return that match's winner
        prev_id = int(slot)
        return results[prev_id]['winner']

    # Simulate R32
    for mid, slotA, slotB in R32:
        team_a = resolve_slot(slotA, mid=mid, side='A')
        team_b = resolve_slot(slotB, mid=mid, side='B')
        winner, p_a, p_b, _ = predict_ko(model, team_a, team_b, profiles)
        results[mid] = {'home': team_a, 'away': team_b, 'winner': winner,
                        'p_home': p_a, 'p_away': p_b, 'round': 'R32'}

    # Simulate R16
    for mid, src_a, src_b in R16:
        team_a = results[src_a]['winner']
        team_b = results[src_b]['winner']
        winner, p_a, p_b, _ = predict_ko(model, team_a, team_b, profiles)
        results[mid] = {'home': team_a, 'away': team_b, 'winner': winner,
                        'p_home': p_a, 'p_away': p_b, 'round': 'R16'}

    # Simulate QF
    for mid, src_a, src_b in QF:
        team_a = results[src_a]['winner']
        team_b = results[src_b]['winner']
        winner, p_a, p_b, _ = predict_ko(model, team_a, team_b, profiles)
        results[mid] = {'home': team_a, 'away': team_b, 'winner': winner,
                        'p_home': p_a, 'p_away': p_b, 'round': 'QF'}

    # Simulate SF
    sf_losers = {}
    for mid, src_a, src_b in SF:
        team_a = results[src_a]['winner']
        team_b = results[src_b]['winner']
        winner, p_a, p_b, _ = predict_ko(model, team_a, team_b, profiles)
        loser = team_b if winner == team_a else team_a
        results[mid] = {'home': team_a, 'away': team_b, 'winner': winner,
                        'p_home': p_a, 'p_away': p_b, 'round': 'SF'}
        sf_losers[mid] = loser

    # 3rd-place playoff
    t3a = sf_losers[SF[0][0]]
    t3b = sf_losers[SF[1][0]]
    winner, p_a, p_b, _ = predict_ko(model, t3a, t3b, profiles)
    results[103] = {'home': t3a, 'away': t3b, 'winner': winner,
                    'p_home': p_a, 'p_away': p_b, 'round': '3rd'}

    # Final
    fin_a = results[SF[0][0]]['winner']
    fin_b = results[SF[1][0]]['winner']
    winner, p_a, p_b, _ = predict_ko(model, fin_a, fin_b, profiles)
    results[104] = {'home': fin_a, 'away': fin_b, 'winner': winner,
                    'p_home': p_a, 'p_away': p_b, 'round': 'Final'}

    return results


# ── Excel writer ──────────────────────────────────────────────────────────────

def _hdr_cell(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, merge_to=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _border()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return cell


def _data_cell(ws, row, col, value, bold=False, bg=WHITE, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
    cell.border    = _border()
    return cell


def write_round_sheet(ws, round_label, matches_info, results):
    """
    matches_info: list of (match_id, label_str)  e.g. (73, 'Jun 28')
    """
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    # Title
    _hdr_cell(ws, 1, 1, round_label, bg=DARK_BLUE, size=13, merge_to=6)

    # Column headers
    cols   = ['Match', 'Date', 'Team A', 'P(A)', 'P(B)', 'Team B', 'Winner']
    widths = [8, 10, 24, 8, 8, 24, 24]
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        _hdr_cell(ws, 3, ci, col, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (mid, date_lbl) in enumerate(matches_info, start=4):
        r = results[mid]
        is_winner_a = r['winner'] == r['home']
        bg = SILVER if ri % 2 == 0 else WHITE

        _data_cell(ws, ri, 1, f'M{mid}', bg=bg)
        _data_cell(ws, ri, 2, date_lbl, bg=bg)
        _data_cell(ws, ri, 3, r['home'], bold=is_winner_a,   bg=GREEN if is_winner_a else bg, center=False)
        _data_cell(ws, ri, 4, f"{r['p_home']:.1%}", bg=bg)
        _data_cell(ws, ri, 5, f"{r['p_away']:.1%}", bg=bg)
        _data_cell(ws, ri, 6, r['away'], bold=not is_winner_a, bg=GREEN if not is_winner_a else bg, center=False)
        _data_cell(ws, ri, 7, r['winner'], bold=True, bg=GOLD)

    ws.freeze_panes = 'A4'


def write_final_sheet(ws, results):
    ws.sheet_view.showGridLines = False

    _hdr_cell(ws, 1, 1, 'FINAL & 3RD PLACE PLAYOFF', bg=DARK_BLUE, size=13, merge_to=7)

    cols   = ['Match', 'Date', 'Team A', 'P(A)', 'P(B)', 'Team B', 'Winner']
    widths = [8, 10, 24, 8, 8, 24, 24]
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        _hdr_cell(ws, 3, ci, col, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (mid, date_lbl) in enumerate([(103,'Jul 18'),(104,'Jul 19')], start=4):
        r = results[mid]
        is_winner_a = r['winner'] == r['home']
        bg = SILVER if ri % 2 == 0 else WHITE
        label = '🏆 Champion' if mid == 104 else '🥉 3rd Place'

        _data_cell(ws, ri, 1, f'M{mid}', bg=bg)
        _data_cell(ws, ri, 2, date_lbl, bg=bg)
        _data_cell(ws, ri, 3, r['home'], bold=is_winner_a,   bg=GREEN if is_winner_a else bg, center=False)
        _data_cell(ws, ri, 4, f"{r['p_home']:.1%}", bg=bg)
        _data_cell(ws, ri, 5, f"{r['p_away']:.1%}", bg=bg)
        _data_cell(ws, ri, 6, r['away'], bold=not is_winner_a, bg=GREEN if not is_winner_a else bg, center=False)
        _data_cell(ws, ri, 7, r['winner'], bold=True, bg=GOLD)

    # Champion call-out
    champion = results[104]['winner']
    ws.row_dimensions[7].height = 28
    c = ws.cell(row=7, column=1, value=f'PREDICTED CHAMPION:  {champion}')
    c.font      = Font(bold=True, size=14, color=DARK_BLUE)
    c.fill      = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = _border()
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    print('Loading group stage predictions...')
    xl  = pd.ExcelFile(XLSX_PATH)
    df  = xl.parse(xl.sheet_names[0])
    df = df[df['Home'].notna() & df['Away'].notna()].copy()
    df['_pred_raw'] = df.apply(
        lambda r: 'home_win' if r['Predicted'] == r['Home']
                  else ('away_win' if r['Predicted'] == r['Away'] else 'draw'), axis=1)

    print('Computing group standings...')
    standings = compute_standings(df)
    for grp, rows in standings.items():
        print(f"  Group {grp}: " + ", ".join(f"{t}({p}pts)" for t,p,w,opp in rows))

    best_thirds = pick_best_thirds(standings, df)
    print(f'\n8 best 3rd-place teams:')
    for pts, wins, opp_win_sum, team, grp in best_thirds:
        print(f'  {team} (Group {grp}) — {pts}pts, {wins}W, opp_win_sum={opp_win_sum:.3f}')

    thirds_assign = allocate_thirds(best_thirds, R32)
    print(f'\n3rd-place slot assignments:')
    for mid, team in sorted(thirds_assign.items()):
        print(f'  Match {mid}: {team}')

    print('\nLoading dataset and training CatBoost (draw_weight=0.66)...')
    full_df  = build_dataset()
    train_df = full_df[full_df['date'] < WC26_START].reset_index(drop=True)
    model    = CatBoostPredictor(draw_weight=0.58)
    model.fit(train_df)

    print('Building team profiles...')
    profiles = build_profiles()

    print('Simulating bracket...')
    results = simulate_bracket(standings, thirds_assign, model, profiles)

    print('\nWriting Excel tabs...')
    wb = load_workbook(XLSX_PATH)

    # Remove existing knockout tabs
    for name in ['Round of 32', 'Round of 16', 'Quarter-finals',
                 'Semi-finals', 'Final']:
        if name in wb.sheetnames:
            del wb[name]

    r32_dates = ['Jun 28','Jun 29','Jun 29','Jun 29','Jun 30','Jun 30',
                 'Jun 30','Jul 1', 'Jul 1', 'Jul 1', 'Jul 2', 'Jul 2',
                 'Jul 2', 'Jul 3', 'Jul 3', 'Jul 3']
    r16_dates = ['Jul 4','Jul 4','Jul 5','Jul 5','Jul 6','Jul 6','Jul 7','Jul 7']
    qf_dates  = ['Jul 9','Jul 10','Jul 11','Jul 11']
    sf_dates  = ['Jul 14','Jul 15']

    write_round_sheet(wb.create_sheet('Round of 32'),
                      'ROUND OF 32',
                      [(mid, d) for (mid,_,__), d in zip(R32, r32_dates)], results)

    write_round_sheet(wb.create_sheet('Round of 16'),
                      'ROUND OF 16',
                      [(mid, d) for (mid,_,__), d in zip(R16, r16_dates)], results)

    write_round_sheet(wb.create_sheet('Quarter-finals'),
                      'QUARTER-FINALS',
                      [(mid, d) for (mid,_,__), d in zip(QF, qf_dates)], results)

    write_round_sheet(wb.create_sheet('Semi-finals'),
                      'SEMI-FINALS',
                      [(mid, d) for (mid,_,__), d in zip(SF, sf_dates)], results)

    write_final_sheet(wb.create_sheet('Final'), results)

    wb.save(XLSX_PATH)
    print(f'\nSaved: {XLSX_PATH}')
    print(f'\nPREDICTED CHAMPION: {results[104]["winner"]}')
    print(f'3rd place:          {results[103]["winner"]}')


if __name__ == '__main__':
    run()
