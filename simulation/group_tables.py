"""
Adds group table tabs (A–L) to simulation/output/wc_26_sim.xlsx.

Each tab shows:
  - Group matches with model predictions and probabilities
  - Predicted standings (Pos, Team, MP, W, D, L, Pts)

Run from the project root:
    python -m simulation.group_tables
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = 'simulation/output/wc_26_sim.xlsx'

GROUPS = {
    'A': ['Mexico', 'South Africa', 'South Korea', 'Czech Republic'],
    'B': ['Canada', 'Bosnia and Herzegovina', 'Qatar', 'Switzerland'],
    'C': ['Brazil', 'Morocco', 'Haiti', 'Scotland'],
    'D': ['United States', 'Paraguay', 'Australia', 'Turkey'],
    'E': ['Germany', 'Curaçao', 'Ivory Coast', 'Ecuador'],
    'F': ['Netherlands', 'Japan', 'Sweden', 'Tunisia'],
    'G': ['Belgium', 'Egypt', 'Iran', 'New Zealand'],
    'H': ['Spain', 'Cape Verde', 'Saudi Arabia', 'Uruguay'],
    'I': ['France', 'Senegal', 'Iraq', 'Norway'],
    'J': ['Argentina', 'Algeria', 'Austria', 'Jordan'],
    'K': ['Portugal', 'DR Congo', 'Uzbekistan', 'Colombia'],
    'L': ['England', 'Croatia', 'Ghana', 'Panama'],
}

# Colours
DARK_BLUE  = 'FF1F4E79'
MID_BLUE   = 'FF2E75B6'
LIGHT_BLUE = 'FFD9E1F2'
WHITE      = 'FFFFFFFF'
GRAY       = 'FFF2F2F2'
GREEN      = 'FFE2EFDA'
GOLD       = 'FFFFF2CC'


def _side():
    return Side(style='thin', color='FFBFBFBF')

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _compute_standings(matches, teams):
    """Compute predicted standings from a list of match dicts."""
    stats = {t: {'MP': 0, 'W': 0, 'D': 0, 'L': 0, 'Pts': 0} for t in teams}

    for m in matches:
        h, a, pred = m['Home'], m['Away'], m['_pred_raw']
        stats[h]['MP'] += 1
        stats[a]['MP'] += 1
        if pred == 'home_win':
            stats[h]['W'] += 1;  stats[h]['Pts'] += 3
            stats[a]['L'] += 1
        elif pred == 'draw':
            stats[h]['D'] += 1;  stats[h]['Pts'] += 1
            stats[a]['D'] += 1;  stats[a]['Pts'] += 1
        else:
            stats[a]['W'] += 1;  stats[a]['Pts'] += 3
            stats[h]['L'] += 1

    table = [{'Team': t, **stats[t]} for t in teams]
    table.sort(key=lambda r: (-r['Pts'], -r['W']))
    for i, row in enumerate(table, 1):
        row['Pos'] = i
    return table


def _write_group_sheet(ws, group_letter, teams, matches):
    ws.sheet_view.showGridLines = False

    # ── Group header ──────────────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value=f'GROUP {group_letter}')
    title_cell.font      = Font(bold=True, color=WHITE, size=13)
    title_cell.fill      = PatternFill('solid', fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # ── Matches section ───────────────────────────────────────────────────
    match_hdr_row = 3
    match_cols    = ['Date', 'Home', 'Away', 'Predicted', 'P(HW)', 'P(D)', 'P(AW)', 'Confidence']
    col_widths    = [12, 22, 22, 22, 7, 7, 7, 12]

    # Column header
    for ci, (col, w) in enumerate(zip(match_cols, col_widths), start=1):
        cell = ws.cell(row=match_hdr_row, column=ci, value=col)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left')
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Match rows
    for ri, m in enumerate(matches, start=match_hdr_row + 1):
        pred_raw  = m['_pred_raw']
        pred_lbl  = m['Predicted']
        row_fill  = PatternFill('solid', fgColor=GRAY if ri % 2 == 0 else WHITE)

        vals = [m['Date'], m['Home'], m['Away'], pred_lbl,
                f"{m['P(HW)']:.3f}", f"{m['P(D)']:.3f}", f"{m['P(AW)']:.3f}", m['Confidence']]

        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal='center' if ci > 3 else 'left')
            cell.border    = _border()
            cell.fill      = row_fill

        # Highlight predicted winner cell
        pred_cell = ws.cell(row=ri, column=4)
        pred_cell.font = Font(bold=True)

    # ── Standings section ─────────────────────────────────────────────────
    standings = _compute_standings(matches, teams)
    stand_hdr_row = match_hdr_row + len(matches) + 3

    ws.cell(row=stand_hdr_row - 1, column=1,
            value='PREDICTED STANDINGS').font = Font(bold=True, color=DARK_BLUE, size=11)

    stand_cols   = ['Pos', 'Team', 'MP', 'W', 'D', 'L', 'Pts']
    stand_widths = [5, 22, 5, 5, 5, 5, 5]

    for ci, (col, w) in enumerate(zip(stand_cols, stand_widths), start=1):
        cell = ws.cell(row=stand_hdr_row, column=ci, value=col)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal='center' if ci != 2 else 'left')
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(standings, start=stand_hdr_row + 1):
        # Top 2 advance → light green; bottom 2 → white
        advancing = ri <= stand_hdr_row + 2
        row_fill  = PatternFill('solid', fgColor=GREEN if advancing else WHITE)
        is_first  = (ri == stand_hdr_row + 1)

        vals = [row['Pos'], row['Team'], row['MP'], row['W'], row['D'], row['L'], row['Pts']]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal='center' if ci != 2 else 'left')
            cell.border    = _border()
            cell.fill      = row_fill
            if is_first:
                cell.font = Font(bold=True)

    ws.freeze_panes = 'A2'


def _write_best_thirds_sheet(ws, all_thirds):
    """
    all_thirds: list of dicts sorted by rank, each with:
      pos, grp, team, mp, w, d, l, pts  (gf/ga/gd are 0 — no score model)
    Top 8 advance.
    """
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    # Title
    title = ws.cell(row=1, column=1, value='RANKING OF THIRD-PLACED TEAMS')
    title.font      = Font(bold=True, color=WHITE, size=13)
    title.fill      = PatternFill('solid', fgColor=DARK_BLUE)
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    # Column headers — match Wikipedia layout
    cols   = ['Pos', 'Grp', 'Team', 'Pld', 'W', 'D', 'L', 'GF', 'GA', 'GD', 'Pts', 'Qualification']
    widths = [5, 5, 24, 5, 5, 5, 5, 5, 5, 5, 5, 28]

    for ci, (col, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=3, column=ci, value=col)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal='center' if ci != 3 else 'left')
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, t in enumerate(all_thirds, start=4):
        advancing = t['pos'] <= 8
        row_fill  = PatternFill('solid', fgColor=GREEN if advancing else WHITE)
        qual_text = 'Advance to knockout stage' if advancing else ''

        # Merge Qualification cell vertically for the 8-team block on first advancing row
        vals = [t['pos'], t['grp'], t['team'],
                t['mp'], t['w'], t['d'], t['l'],
                '-', '-', '-',   # GF, GA, GD — no score predictions
                t['pts'], qual_text]

        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.border    = _border()
            cell.alignment = Alignment(horizontal='center' if ci not in (3, 12) else 'left',
                                       vertical='center')
            if ci == 11:  # Pts — bold
                cell.font = Font(bold=True)

    ws.freeze_panes = 'A4'


def _opp_win_prob_sum(team, df):
    """Sum of opponent win probabilities across a team's group matches (lower = better)."""
    total = 0.0
    for _, row in df.iterrows():
        if row['Home'] == team:
            total += row['P(AW)']
        elif row['Away'] == team:
            total += row['P(HW)']
    return total


def _compute_all_thirds(df, elo_map):
    """
    Compute standing for every group, extract 3rd-place teams,
    rank by Pts -> W -> opponent win prob sum (ascending). Returns list of dicts sorted best to worst.
    """
    all_groups_standings = {}
    for grp, teams in GROUPS.items():
        mask    = df['Home'].isin(teams) & df['Away'].isin(teams)
        matches = df[mask].to_dict('records')
        table   = _compute_standings(matches, teams)
        all_groups_standings[grp] = table

    thirds = []
    for grp, table in all_groups_standings.items():
        t    = table[2]   # 3rd place row
        team = t['Team']
        thirds.append({
            'grp':         grp,
            'team':        team,
            'mp':          t['MP'],
            'w':           t['W'],
            'd':           t['D'],
            'l':           t['L'],
            'pts':         t['Pts'],
            'opp_win_sum': _opp_win_prob_sum(team, df),
        })

    thirds.sort(key=lambda x: (-x['pts'], -x['w'], x['opp_win_sum']))
    for i, t in enumerate(thirds, 1):
        t['pos'] = i
    return thirds


def run():
    wb = load_workbook(XLSX_PATH)

    # Read predictions from the main sheet
    xl  = pd.ExcelFile(XLSX_PATH)
    df  = xl.parse(xl.sheet_names[0])
    df  = df[df['Home'].notna() & df['Away'].notna()].copy()

    def infer_pred_raw(row):
        if row['Predicted'] == row['Home']:   return 'home_win'
        if row['Predicted'] == row['Away']:   return 'away_win'
        return 'draw'

    df['_pred_raw'] = df.apply(infer_pred_raw, axis=1)

    elo_map = {}  # no longer used for tiebreaking

    # Remove existing tabs so we don't duplicate
    for name in list(GROUPS.keys()):
        if f'Group {name}' in wb.sheetnames:
            del wb[f'Group {name}']
    if 'BEST 3rds' in wb.sheetnames:
        del wb['BEST 3rds']

    for letter, teams in GROUPS.items():
        mask    = df['Home'].isin(teams) & df['Away'].isin(teams)
        g_df    = df[mask].sort_values('Date').reset_index(drop=True)
        matches = g_df.to_dict('records')

        ws = wb.create_sheet(title=f'Group {letter}')
        _write_group_sheet(ws, letter, teams, matches)
        print(f'  Group {letter}: {len(matches)} matches')

    # Best 3rd-place tab
    all_thirds = _compute_all_thirds(df, elo_map)
    ws3 = wb.create_sheet(title='BEST 3rds')
    _write_best_thirds_sheet(ws3, all_thirds)
    print(f'  BEST 3rds tab written ({len(all_thirds)} teams)')

    wb.save(XLSX_PATH)
    print(f'\nSaved: {XLSX_PATH}')


if __name__ == '__main__':
    run()
