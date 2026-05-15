"""
Reads per-match results from backtest/output/*.xlsx and computes average
metrics (Accuracy, Log-Loss, RPS) per model across all WC editions.

Output: backtest/avg_performance.xlsx  — three stacked tables (Accuracy / Log-Loss / RPS),
        one row per model, columns: Model | WC22 | WC18 | WC14 | WC10 | WC06 | Avg

Run from the project root:
    python -m backtest.avg_performance
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = 'backtest/output'
SHEETS     = ['WC 22', 'WC 18', 'WC 14', 'WC 10', 'WC 06']
EDITIONS   = ['WC22',  'WC18',  'WC14',  'WC10',  'WC06']

# Colours
HDR_SECTION = 'FF1F4E79'   # dark blue — section title (ACCURACY / LOG-LOSS / RPS)
HDR_COL     = 'FFD9E1F2'   # light blue — column headers row
WHITE       = 'FFFFFFFF'
DARK_FONT   = 'FF1F4E79'

def _side():
    return Side(style='thin', color='FFBFBFBF')

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def read_model_results(xlsx_path):
    xl = pd.ExcelFile(xlsx_path)
    edition_dfs = {}
    for sheet, ed in zip(SHEETS, EDITIONS):
        if sheet not in xl.sheet_names:
            continue
        df = xl.parse(sheet, header=0)
        df = df[pd.to_datetime(df['Date'], errors='coerce').notna()].copy()
        df['RPS']      = pd.to_numeric(df['RPS'],      errors='coerce')
        df['Log-Loss'] = pd.to_numeric(df['Log-Loss'], errors='coerce')
        edition_dfs[ed] = df
    return edition_dfs


def compute_metrics(df):
    return (df['Result'] == 'OK').mean(), df['Log-Loss'].mean(), df['RPS'].mean()


def build_table(records, metric_key, fmt_fn):
    """Return list of row-dicts sorted by Avg (best first)."""
    rows = []
    for rec in records:
        vals = [rec[metric_key][ed] for ed in EDITIONS]
        avg  = sum(vals) / len(vals)
        rows.append({'Model': rec['model'], **{ed: v for ed, v in zip(EDITIONS, vals)}, 'Avg': avg})
    # sort: accuracy → higher better; log_loss / rps → lower better
    reverse = (metric_key == 'accuracy')
    rows.sort(key=lambda r: r['Avg'], reverse=reverse)
    return rows


def write_section(ws, title, table_rows, fmt_fn, start_row):
    cols = ['Model'] + EDITIONS + ['Avg']
    col_widths = [16, 8, 8, 8, 8, 8, 8]

    # Section title row
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font      = Font(bold=True, color=WHITE, size=12)
    title_cell.fill      = PatternFill('solid', fgColor=HDR_SECTION)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[start_row].height = 18
    # merge across all columns
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cols))

    # Column header row
    hdr_row = start_row + 1
    for ci, (col, width) in enumerate(zip(cols, col_widths), start=1):
        cell = ws.cell(row=hdr_row, column=ci, value=col)
        cell.font      = Font(bold=True, color=DARK_FONT)
        cell.fill      = PatternFill('solid', fgColor=HDR_COL)
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left')
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, row in enumerate(table_rows, start=hdr_row + 1):
        for ci, col in enumerate(cols, start=1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci, value=fmt_fn(val) if ci > 1 else val)
            cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left')
            cell.border    = _border()
            # Bold the Avg column
            if ci == len(cols):
                cell.font = Font(bold=True)

    return start_row + 1 + len(table_rows) + 2   # next free row (with a blank gap)


def run():
    xlsx_files = sorted(glob.glob(os.path.join(OUTPUT_DIR, '*.xlsx')))
    if not xlsx_files:
        print(f'No .xlsx files found in {OUTPUT_DIR}/')
        return

    records = []
    for xlsx_path in xlsx_files:
        model_name  = os.path.splitext(os.path.basename(xlsx_path))[0]
        edition_dfs = read_model_results(xlsx_path)
        if not edition_dfs:
            continue
        acc_vals = {}
        ll_vals  = {}
        rps_vals = {}
        for ed, df in edition_dfs.items():
            acc, ll, rps = compute_metrics(df)
            acc_vals[ed] = round(acc, 4)
            ll_vals[ed]  = round(ll,  4)
            rps_vals[ed] = round(rps, 4)
        records.append({'model': model_name, 'accuracy': acc_vals, 'log_loss': ll_vals, 'rps': rps_vals})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    cur_row = 1
    cur_row = write_section(ws, 'ACCURACY',  build_table(records, 'accuracy', lambda v: f'{v:.1%}'),   lambda v: f'{v:.1%}',   cur_row)
    cur_row = write_section(ws, 'LOG-LOSS',  build_table(records, 'log_loss', lambda v: f'{v:.4f}'),   lambda v: f'{v:.4f}',   cur_row)
    cur_row = write_section(ws, 'RPS',       build_table(records, 'rps',      lambda v: f'{v:.4f}'),   lambda v: f'{v:.4f}',   cur_row)

    out_path = 'backtest/avg_performance.xlsx'
    wb.save(out_path)
    print(f'Saved: {out_path}')

    # Console summary
    acc_table = build_table(records, 'accuracy', None)
    print(f"\n{'ACCURACY'}")
    print(f"  {'Model':<16}  " + "  ".join(f"{ed:>6}" for ed in EDITIONS) + f"  {'Avg':>6}")
    print("  " + "-" * 70)
    for r in acc_table:
        print(f"  {r['Model']:<16}  " + "  ".join(f"{r[ed]:>6.1%}" for ed in EDITIONS) + f"  {r['Avg']:>6.1%}")

    rps_table = build_table(records, 'rps', None)
    print(f"\n{'RPS'}")
    print(f"  {'Model':<16}  " + "  ".join(f"{ed:>7}" for ed in EDITIONS) + f"  {'Avg':>7}")
    print("  " + "-" * 70)
    for r in rps_table:
        print(f"  {r['Model']:<16}  " + "  ".join(f"{r[ed]:>7.4f}" for ed in EDITIONS) + f"  {r['Avg']:>7.4f}")


if __name__ == '__main__':
    run()
