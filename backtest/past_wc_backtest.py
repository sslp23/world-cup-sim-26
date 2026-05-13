"""
Cross-WC backtest — evaluates all models on each past World Cup edition.

For each WC edition the dataset in data/past_wc/wc{year}/ is used:
  - Training set : all ranked matches before the WC start date
  - Test set     : the 64 WC matches of that edition

Models evaluated:
  Dixon-Coles | XGBoost | CatBoost | Ordered Logit | ML-Poisson | Ensemble

Output: backtest/output/{model}.xlsx
  Each file has one sheet per WC edition (WC 22, WC 18, ..., WC 06).
  Each sheet contains the per-match results table + aggregate metrics block.

Run from the project root:
    python -m backtest.past_wc_backtest
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.dixon_coles.model   import DixonColes
from models.xgboost.model       import XGBoostPredictor
from models.catboost.model      import CatBoostPredictor
from models.ordered_logit.model import OrderedLogitPredictor
from models.ml_poisson.model    import MLPoissonModel

EDITIONS = [
    (2022, '2022-11-20', '2022-12-18'),
    (2018, '2018-06-14', '2018-07-15'),
    (2014, '2014-06-12', '2014-07-13'),
    (2010, '2010-06-11', '2010-07-11'),
    (2006, '2006-06-09', '2006-07-09'),
]

OUTCOME_ORDER = ['home_win', 'draw', 'away_win']

STAGE_MAP = [
    ('Group stage',    lambda d, s, e: d <= s + pd.Timedelta(days=13 if s.month == 11 else 21)),
    ('Round of 16',   None),
    ('Quarter-finals',None),
    ('Semi-finals',   None),
    ('Final / 3rd',   None),
]

# Stage date ranges per year (inclusive)
STAGE_DATES = {
    2022: [
        ('Group stage',    '2022-11-20', '2022-12-02'),
        ('Round of 16',    '2022-12-03', '2022-12-06'),
        ('Quarter-finals', '2022-12-09', '2022-12-10'),
        ('Semi-finals',    '2022-12-13', '2022-12-14'),
        ('Final / 3rd',    '2022-12-17', '2022-12-18'),
    ],
    2018: [
        ('Group stage',    '2018-06-14', '2018-06-28'),
        ('Round of 16',    '2018-06-30', '2018-07-03'),
        ('Quarter-finals', '2018-07-06', '2018-07-07'),
        ('Semi-finals',    '2018-07-10', '2018-07-11'),
        ('Final / 3rd',    '2018-07-14', '2018-07-15'),
    ],
    2014: [
        ('Group stage',    '2014-06-12', '2014-06-26'),
        ('Round of 16',    '2014-06-28', '2014-07-01'),
        ('Quarter-finals', '2014-07-04', '2014-07-05'),
        ('Semi-finals',    '2014-07-08', '2014-07-09'),
        ('Final / 3rd',    '2014-07-12', '2014-07-13'),
    ],
    2010: [
        ('Group stage',    '2010-06-11', '2010-06-25'),
        ('Round of 16',    '2010-06-26', '2010-06-29'),
        ('Quarter-finals', '2010-07-02', '2010-07-03'),
        ('Semi-finals',    '2010-07-06', '2010-07-07'),
        ('Final / 3rd',    '2010-07-10', '2010-07-11'),
    ],
    2006: [
        ('Group stage',    '2006-06-09', '2006-06-23'),
        ('Round of 16',    '2006-06-24', '2006-06-27'),
        ('Quarter-finals', '2006-06-30', '2006-07-01'),
        ('Semi-finals',    '2006-07-04', '2006-07-05'),
        ('Final / 3rd',    '2006-07-08', '2006-07-09'),
    ],
}


# ── Metrics ────────────────────────────────────────────────────────────────────

def actual_outcome(home_score, away_score):
    if home_score > away_score:    return 'home_win'
    elif home_score == away_score: return 'draw'
    return 'away_win'


def rps(probs, outcome):
    actual_vec = [1.0 if o == outcome else 0.0 for o in OUTCOME_ORDER]
    pred_cum   = np.cumsum([probs[o] for o in OUTCOME_ORDER])
    actual_cum = np.cumsum(actual_vec)
    return float(np.sum((pred_cum - actual_cum) ** 2) / (len(OUTCOME_ORDER) - 1))


def log_loss_single(probs, outcome):
    return -np.log(probs[outcome] + 1e-15)


def avg_probs(*prob_dicts):
    result = {o: 0.0 for o in OUTCOME_ORDER}
    for p in prob_dicts:
        for o in OUTCOME_ORDER:
            result[o] += p[o] / len(prob_dicts)
    return result


def predict_all(get_probs_fn, wc_df, desc=''):
    """Run model on all WC matches, return list of row dicts."""
    rows = []
    for _, match in tqdm(wc_df.iterrows(), total=len(wc_df), desc=desc, leave=False):
        probs   = get_probs_fn(match)
        outcome = actual_outcome(match['home_score'], match['away_score'])
        pred    = max(probs, key=probs.get)
        rows.append({
            'Date':       str(match['date'].date()),
            'Home':       match['home_team'],
            'Away':       match['away_team'],
            'Score':      f"{int(match['home_score'])}-{int(match['away_score'])}",
            'Actual':     outcome,
            'Predicted':  pred,
            'Result':     'OK' if pred == outcome else 'FAIL',
            'P(HW)':      round(probs['home_win'], 3),
            'P(D)':       round(probs['draw'], 3),
            'P(AW)':      round(probs['away_win'], 3),
            'RPS':        round(rps(probs, outcome), 4),
            'Log-Loss':   round(log_loss_single(probs, outcome), 4),
            '_correct':   int(pred == outcome),
        })
    return rows


# ── Excel writer ───────────────────────────────────────────────────────────────

GREEN  = 'FFD6F5DC'
RED    = 'FFFFC7CE'
HEADER = 'FF1F4E79'
SUBHDR = 'FFD9E1F2'
GRAY   = 'FFF2F2F2'

def _border():
    s = Side(style='thin', color='FFBFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(ws, rows, year):
    """Write per-match rows + aggregate metrics to a worksheet."""
    cols = ['Date', 'Home', 'Away', 'Score', 'Actual', 'Predicted', 'Result',
            'P(HW)', 'P(D)', 'P(AW)', 'RPS', 'Log-Loss']
    col_widths = [12, 24, 24, 7, 11, 11, 7, 7, 7, 7, 8, 9]

    # Header row
    header_font  = Font(bold=True, color='FFFFFFFF')
    header_fill  = PatternFill('solid', fgColor=HEADER)
    header_align = Alignment(horizontal='center')

    for ci, (col, width) in enumerate(zip(cols, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, row in enumerate(rows, start=2):
        is_ok   = row['Result'] == 'OK'
        fill    = PatternFill('solid', fgColor=GREEN if is_ok else RED)
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci, value=row[col])
            cell.border = _border()
            cell.alignment = Alignment(horizontal='center' if ci > 3 else 'left')
            if col == 'Result':
                cell.fill = fill

    # Blank separator
    sep_row = len(rows) + 3

    # Aggregate metrics
    n       = len(rows)
    correct = sum(r['_correct'] for r in rows)
    acc     = correct / n
    mean_ll = sum(r['Log-Loss'] for r in rows) / n
    mean_rps= sum(r['RPS'] for r in rows) / n
    base_rps= sum(rps({'home_win':1/3,'draw':1/3,'away_win':1/3},
                      r['Actual']) for r in rows) / n
    base_ll = -np.log(1/3)

    subhdr_fill = PatternFill('solid', fgColor=SUBHDR)
    bold = Font(bold=True)

    def metric_row(r, label, value):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = bold
        cell.fill = PatternFill('solid', fgColor=GRAY)
        ws.cell(row=r, column=2, value=value)

    ws.cell(row=sep_row - 1, column=1, value='AGGREGATE METRICS').font = Font(bold=True, color='FF1F4E79', size=11)
    metric_row(sep_row,     'Accuracy',        f'{acc:.1%}  ({correct}/{n})')
    metric_row(sep_row + 1, 'Log-Loss',        f'{mean_ll:.4f}  (baseline: {base_ll:.4f})')
    metric_row(sep_row + 2, 'RPS',             f'{mean_rps:.4f}  (baseline: {base_rps:.4f})')

    # By stage
    stage_row = sep_row + 4
    ws.cell(row=stage_row - 1, column=1, value='ACCURACY BY STAGE').font = Font(bold=True, color='FF1F4E79', size=11)
    for stage, s_start, s_end in STAGE_DATES[year]:
        s_rows = [r for r in rows
                  if pd.Timestamp(s_start) <= pd.Timestamp(r['Date']) <= pd.Timestamp(s_end)]
        if s_rows:
            s_correct = sum(r['_correct'] for r in s_rows)
            s_n       = len(s_rows)
            metric_row(stage_row, stage, f'{s_correct/s_n:.1%}  ({s_correct}/{s_n})')
            stage_row += 1

    ws.freeze_panes = 'A2'


def save_excel(model_name, sheets_data, out_dir):
    """sheets_data: list of (year, rows)"""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for year, rows in sheets_data:
        tab_name = f'WC {str(year)[2:]}'
        ws = wb.create_sheet(title=tab_name)
        write_sheet(ws, rows, year)
        print(f'    Written sheet: {tab_name}  ({len(rows)} matches)')

    path = os.path.join(out_dir, f'{model_name}.xlsx')
    wb.save(path)
    print(f'  Saved: {path}')


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    out_dir = 'backtest/output'
    os.makedirs(out_dir, exist_ok=True)

    # Store per-model results: {model_name: [(year, rows), ...]}
    model_sheets = {
        'Dixon-Coles': [],
        'XGBoost':     [],
        'CatBoost':    [],
        'Ord_Logit':   [],
        'ML-Poisson':  [],
        'Ensemble':    [],
    }

    all_agg = {}   # {year: {model: metrics}} for summary table

    for year, wc_start, wc_end in tqdm(EDITIONS, desc='WC editions'):
        wc_start_ts = pd.Timestamp(wc_start)
        wc_end_ts   = pd.Timestamp(wc_end)

        print(f'\n{"="*60}')
        print(f'  WC {year}  ({wc_start} -> {wc_end})')
        print(f'{"="*60}')

        data_path = f'data/past_wc/wc{year}/ranked_database_with_features.csv'
        full_df = pd.read_csv(data_path)
        full_df['date'] = pd.to_datetime(full_df['date'])

        train_df = full_df[full_df['date'] < wc_start_ts].reset_index(drop=True)
        wc_df    = full_df[
            (full_df['tournament'] == 'FIFA World Cup') &
            (full_df['date'] >= wc_start_ts) &
            (full_df['date'] <= wc_end_ts)
        ].copy().reset_index(drop=True)

        print(f'  Training rows: {len(train_df)}  |  WC matches: {len(wc_df)}')

        # Train
        dc  = DixonColes(xi=0.0005)
        dc.fit(train_df, ref_date=wc_start_ts)

        xgb = XGBoostPredictor(draw_weight=0.65)
        xgb.fit(train_df)

        cb  = CatBoostPredictor(draw_weight=0.72)
        cb.fit(train_df)

        ol  = OrderedLogitPredictor()
        ol.fit(train_df)

        mlp = MLPoissonModel(rho=-0.40)
        mlp.fit(train_df)

        models = {
            'Dixon-Coles': lambda m: dc.predict_outcome_probs(m['home_team'], m['away_team'], neutral=True),
            'XGBoost':     lambda m: xgb.predict_proba_row(m),
            'CatBoost':    lambda m: cb.predict_proba_row(m),
            'Ord_Logit':   lambda m: ol.predict_proba_row(m),
            'ML-Poisson':  lambda m: mlp.predict_proba_row(m),
            'Ensemble':    lambda m: avg_probs(xgb.predict_proba_row(m),
                                               cb.predict_proba_row(m),
                                               mlp.predict_proba_row(m)),
        }

        all_agg[year] = {}
        for name, fn in tqdm(models.items(), desc=f'  WC{year} models', leave=False):
            rows = predict_all(fn, wc_df, desc=f'    {name}')
            model_sheets[name].append((year, rows))
            n       = len(rows)
            correct = sum(r['_correct'] for r in rows)
            mean_ll = sum(r['Log-Loss'] for r in rows) / n
            mean_rps= sum(r['RPS'] for r in rows) / n
            all_agg[year][name] = {'accuracy': correct/n, 'log_loss': mean_ll, 'rps': mean_rps,
                                   'correct': correct, 'n': n}
            print(f'  {name:<14}  acc={correct/n:.1%} ({correct}/{n})  ll={mean_ll:.4f}  rps={mean_rps:.4f}')

    # ── Write Excel files ──────────────────────────────────────────────────────
    print(f'\nWriting Excel files to {out_dir}/ ...')
    for model_name, sheets_data in model_sheets.items():
        save_excel(model_name, sheets_data, out_dir)

    # ── Summary tables ─────────────────────────────────────────────────────────
    model_names = ['Dixon-Coles', 'XGBoost', 'CatBoost', 'Ord_Logit', 'ML-Poisson', 'Ensemble']
    years       = [y for y, _, _ in EDITIONS]

    for metric, label, fmt in [
        ('accuracy', 'ACCURACY', '{:.1%}'),
        ('log_loss', 'LOG-LOSS', '{:.4f}'),
        ('rps',      'RPS',      '{:.4f}'),
    ]:
        print(f'\n  {label}')
        col_w = 12
        header = f"  {'Model':<14}" + "".join(f"  {'WC'+str(y)[2:]:>{col_w}}" for y in years) + f"  {'Avg':>{col_w}}"
        print(header)
        print("  " + "-" * (14 + (col_w + 2) * (len(years) + 1)))
        for name in model_names:
            vals = [all_agg[y][name][metric] for y in years]
            avg  = np.mean(vals)
            row  = f"  {name:<14}"
            for v in vals:
                row += f"  {fmt.format(v):>{col_w}}"
            row += f"  {fmt.format(avg):>{col_w}}"
            print(row)

    print()


if __name__ == '__main__':
    run()
